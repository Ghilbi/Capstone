import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import random
from typing import List, Dict, Tuple, Set
import openpyxl
from datetime import datetime, timedelta
import copy

############################################
#               DATA CLASSES               #
############################################

class Course:
    def __init__(self, code: str, description: str, units: int, time: str, days: str, room: str):
        self.code = code
        self.description = description
        self.units = units  # Added units field
        self.time = time
        self.days = days
        self.room = room
        self.time_obj = datetime.strptime(time.split('-')[0], '%I:%M%p')
        self.end_time_obj = datetime.strptime(time.split('-')[1], '%I:%M%p')
        self.is_lab = '(Lab)' in description
        self.is_lec = '(Lec)' in description
        self.base_code = self.get_base_code()
        self.pair_key = self.get_pair_key()

    def get_base_code(self) -> str:
        return self.code.split('(')[0].strip()
    
    def get_pair_key(self) -> str:
        base_desc = self.description.split('(')[0].strip()
        return f"{self.base_code}_{base_desc}"

    def __lt__(self, other):
        return self.time_obj < other.time_obj

    def clone(self):
        return copy.deepcopy(self)

    def is_restricted_room_course(self) -> bool:
        restricted_codes = {'NSTP1', 'PathFit', 'PATHFit'}
        return any(code in self.code for code in restricted_codes)


class TimeSlot:
    def __init__(self, start_time: str):
        self.start_time = datetime.strptime(start_time, '%I:%M%p')
        self.end_time = self.start_time + timedelta(minutes=80)
        self.start_time_str = start_time
        self.end_time_str = self.end_time.strftime('%I:%M%p')
        self.mwf_used = False
        self.tth_used = False

    def is_available(self, is_mwf: bool) -> bool:
        return not (self.mwf_used if is_mwf else self.tth_used)

    def mark_used(self, is_mwf: bool):
        if is_mwf:
            self.mwf_used = True
        else:
            self.tth_used = True

    def __lt__(self, other):
        return self.start_time < other.start_time


############################################
#        HELPER FUNCTIONS / CONSTANTS      #
############################################

DAY_PATTERN_PAIRS = {
    'MW': 'MWF',    # If lecture is MW, lab must be MWF
    'TTH': 'TTHS'   # If lecture is TTH, lab must be TTHS
}

MAJOR_SUBJECTS = {
    'CC1', 'CC10', 'CC11', 'CC12', 'CC13', 'CC14', 'CC15', 'CC16', 'CC17', 'CC18', 'CC19', 'CC2',
    'CC21', 'CC22', 'CC23', 'CC24', 'CC3', 'CC4', 'CC5', 'CC6', 'CC7', 'CC8', 'CC9', 'CCS10',
    'CCS11', 'CCS12', 'CCS13', 'CCS14', 'CCS15', 'CCS16', 'CCS2', 'CCS3', 'CCS4', 'CCS5', 'CCS6',
    'CCS7', 'CCS8', 'CCS9', 'CDA1', 'CDA10', 'CDA11', 'CDA12', 'CDA2', 'CDA3', 'CDA4', 'CDA5',
    'CDA6', 'CDA7', 'CDA8', 'CDA9', 'CIT1', 'CIT10', 'CIT11', 'CIT12', 'CIT14', 'CIT15', 'CIT16',
    'CIT17', 'CIT18', 'CIT19', 'CIT26', 'CIT22', 'CIT23', 'CIT24', 'CIT25', 'CIT3', 'CIT4', 'CIT5',
    'CIT6', 'CIT7', 'CIT8', 'CIT9', 'CCS1', 'MCC1', 'MCC2', 'MCC3', 'MCC4', 'MCC5', 'MCC6', 'MCC7', 
    'MCC8', 'MM12', 'MMC1', 'MMC11', 'MMC13','MMC14', 'MMC15', 'MMC16', 'MMC17', 'MMC18', 'MMC19', 
    'MMC2', 'MMC3', 'MMC4', 'MMC5', 'MMC6', 'MMC7', 'MMC8', 'MMC9', 'CIT13'
}

LAB_ROOMS = {'M303', 'M304', 'M305', 'M306', 'M307', 'N3001', 'N3002', 'S312'}
MAJOR_ROOMS = {'M301', 'M303', 'M304', 'M305', 'M306', 'M307', 'N3001', 'N3002', 'S312'}
RESTRICTED_ROOMS = {'Gym', 'Aud'}

BASE_TIMES = [
    '7:30am', '8:50am', '10:10am', '11:30am',
    '12:50pm', '2:10pm', '3:30pm', '4:50pm', '6:10pm'
]

def get_available_rooms(course: Course, all_rooms: Set[str], year_level: str, trimester: str) -> List[str]:
    """
    Return a list of possible rooms for the course based on:
      - Whether it's an NSTP or PathFit (restricted specific rooms).
      - If it's a lab course (use only lab rooms).
      - If it's a major subject (use major rooms).
      - Otherwise return all rooms except restricted ones, labs, and major rooms.
    """
    # If it's an NSTP course, return only Auditorium
    if 'NSTP' in course.code:
        return ['Aud']
        
    # If it's a PathFit course, return only Gym
    if 'PathFit' in course.code or 'PATHFit' in course.code:
        return ['Gym']

    # Check if the course code matches any major subject code
    is_major = any(course.code.startswith(major_code) for major_code in MAJOR_SUBJECTS)
    
    # If it's a lab course, only return lab rooms
    if course.is_lab:
        return list(LAB_ROOMS)
    
    # If it's a major subject but not a lab
    if is_major:
        return list(MAJOR_ROOMS)
        
    # For regular courses, return all rooms except restricted (Gym, Aud), lab, and major rooms
    return list(all_rooms - RESTRICTED_ROOMS - LAB_ROOMS - MAJOR_ROOMS)


##########################################
# FIXED get_consecutive_time_slots BELOW #
##########################################
def get_consecutive_time_slots(time_slots: List[TimeSlot], is_mwf: bool, num_slots: int) -> List[List[TimeSlot]]:
    """
    Find available consecutive time slots with strictly enforced gap rules
    (each next slot starts exactly where the previous one ended).
    """
    available_sequences = []
    base_times_ordered = [
        '7:30am', '8:50am', '10:10am', '11:30am',
        '12:50pm', '2:10pm', '3:30pm', '4:50pm', '6:10pm'
    ]
    time_slots_dict = {slot.start_time_str: slot for slot in time_slots}
    
    for i in range(len(base_times_ordered) - num_slots + 1):
        current_sequence = []
        valid_sequence = True
        slots_needed = num_slots
        current_index = i
        last_used_index = None
        
        while slots_needed > 0 and current_index < len(base_times_ordered):
            current_time = base_times_ordered[current_index]
            
            # Check if slot exists and is available
            if current_time in time_slots_dict and time_slots_dict[current_time].is_available(is_mwf):
                # Check gap size if not first slot
                if last_used_index is not None:
                    gap = current_index - last_used_index
                    # allow only consecutive (gap must be == 1)
                    if gap > 1:  
                        valid_sequence = False
                        break
                current_sequence.append(time_slots_dict[current_time])
                last_used_index = current_index
                slots_needed -= 1
            
            current_index += 1
        
        if valid_sequence and len(current_sequence) == num_slots:
            available_sequences.append(current_sequence)
    
    return available_sequences


############################################
#        COURSE SCHEDULING LOGIC           #
############################################

def balance_and_shuffle_courses(courses: List[Course], year_level: str, trimester: str) -> List[Course]:
    """
    Distribute courses among MWF or TTH schedules, taking into account 
    whether a course is lab or lecture (paired) or standalone, 
    trying to keep a balance between MWF and TTH usage.
    """
    # Initialize base time slots
    time_slots = [TimeSlot(t) for t in BASE_TIMES]
    
    # Track room assignments per (base) course code
    course_room_history = {}
    
    # Group courses by pairing (lec+lab) or standalone
    paired_courses = {}
    standalone_courses = []
    
    # Track usage of consecutive slots (so as not to overload them)
    consecutive_slot_count = {}
    
    for course in courses:
        if course.is_lab or course.is_lec:
            paired_courses.setdefault(course.pair_key, []).append(course.clone())
        else:
            standalone_courses.append(course.clone())
    
    all_rooms = set(course.room for course in courses)
    
    # Lists to store pairs and standalone assignments
    mwf_pairs = []       # Will store (lecture, lab) tuples
    tth_pairs = []       # Will store (lecture, lab) tuples
    mwf_standalone = []
    tth_standalone = []

    def can_use_single_pattern(yr_level: str, tri: str) -> bool:
        """
        Decide if we allow the schedule to be unbalanced (a single pattern).
        Example rule: For third-year second or third trimester, 30% chance.
        """
        if yr_level == "Third" and tri in ["Second", "Third"]:
            return random.random() < 0.3
        return False
    
    def should_use_mwf(current_mwf_count: int, current_tth_count: int, 
                       total_c: int, allow_unbalanced: bool) -> bool:
        """
        Decide if the next course/pair should use MWF (True) or TTH (False).
        Balances usage between MWF and TTH unless unbalanced is allowed.
        """
        if allow_unbalanced:
            return random.random() < 0.5
        
        target_ratio = 0.5  
        current_ratio = current_mwf_count / max(1, (current_mwf_count + current_tth_count))
        
        # If we're significantly off balance, force correction
        if current_ratio < 0.4:
            return True
        if current_ratio > 0.6:
            return False
        
        # Otherwise, slightly bias toward balancing but allow randomness
        balance_bias = target_ratio - current_ratio
        random_threshold = 0.5 + (balance_bias * 0.5)
        return random.random() < random_threshold
    
    def select_room(course: Course, available_rooms: List[str]) -> str:
        """
        Select an appropriate room for the course.
        Prefer reusing a room for the same course code if possible.
        """
        if course.is_restricted_room_course():
            if 'NSTP' in course.code:
                return 'Aud'
            if 'PathFit' in course.code or 'PATHFit' in course.code:
                return 'Gym'
        
        base_code = course.base_code
        
        # Reuse a previously used room with 60% chance, if available
        if base_code in course_room_history:
            previously_used = [r for r in available_rooms if r in course_room_history[base_code]]
            if previously_used and random.random() < 0.6:
                return random.choice(previously_used)
        
        # Otherwise, pick a random available room
        selected_room = random.choice(available_rooms)
        
        # Record this room usage for future reference
        course_room_history.setdefault(base_code, set()).add(selected_room)
        
        return selected_room

    # Decide if we can allow unbalanced scheduling
    allow_unbalanced = can_use_single_pattern(year_level, trimester)
    total_course_count = len(paired_courses) + len(standalone_courses)
    
    # Process paired (lecture+lab) courses
    pair_keys = list(paired_courses.keys())
    random.shuffle(pair_keys)
    
    for pair_key in pair_keys:
        pair = paired_courses[pair_key]
        # Sort pair so lecture is first, lab is second
        pair.sort(key=lambda x: 0 if x.is_lec else 1)
        
        # Determine rooms for each course in the pair
        rooms_for_pair = []
        for c in pair:
            avail_rooms = get_available_rooms(c, all_rooms, year_level, trimester)
            if not avail_rooms:
                break
            rooms_for_pair.append(avail_rooms)
        
        # If any course in this pair has no available rooms, skip
        if len(rooms_for_pair) != len(pair):
            continue
        
        try:
            # First, assign lecture time slot and pattern
            lec_course = next(c for c in pair if c.is_lec)
            lab_course = next(c for c in pair if c.is_lab)
        except StopIteration:
            # If either lecture or lab is missing in this pair, skip it
            continue

        
        # Decide if we should use MWF
        current_mwf_count = len(mwf_pairs)*2 + len(mwf_standalone)
        current_tth_count = len(tth_pairs)*2 + len(tth_standalone)
        use_mwf = should_use_mwf(current_mwf_count, current_tth_count, 
                                 total_course_count, allow_unbalanced)
        
        lec_pattern = 'MW' if use_mwf else 'TTH'
        lab_pattern = DAY_PATTERN_PAIRS[lec_pattern]
        
        is_mwf_lec = (lec_pattern == 'MW')
        is_mwf_lab = (lab_pattern == 'MWF')
        
        # Find consecutive slots for 2 classes
        lec_sequences = get_consecutive_time_slots(time_slots, is_mwf_lec, 2)
        if not lec_sequences:
            continue
        
        # Choose one consecutive block
        consecutive_slots = random.choice(lec_sequences)
        
        # Mark the two chosen slots as used: first for lecture, second for lab
        lec_slot = consecutive_slots[0]
        lab_slot = consecutive_slots[1]
        lec_slot.mark_used(is_mwf_lec)
        lab_slot.mark_used(is_mwf_lab)
        
        # Lecture scheduling
        lec_course.time = f"{lec_slot.start_time_str}-{lec_slot.end_time_str}"
        lec_course.days = lec_pattern
        lec_course.room = select_room(lec_course, rooms_for_pair[0])
        lec_course.time_obj = lec_slot.start_time
        lec_course.end_time_obj = lec_slot.end_time
        
        # Lab scheduling (right after lecture)
        lab_course.time = f"{lab_slot.start_time_str}-{lab_slot.end_time_str}"
        lab_course.days = lab_pattern
        lab_course.room = select_room(lab_course, rooms_for_pair[1])
        lab_course.time_obj = lab_slot.start_time
        lab_course.end_time_obj = lab_slot.end_time
        
        # Store the pair in appropriate list
        if is_mwf_lec:
            mwf_pairs.append((lec_course, lab_course))
        else:
            tth_pairs.append((lec_course, lab_course))
    
    # Process standalone courses
    random.shuffle(standalone_courses)
    for course in standalone_courses:
        available_rooms = get_available_rooms(course, all_rooms, year_level, trimester)
        if not available_rooms:
            continue
        
        current_mwf_count = len(mwf_pairs)*2 + len(mwf_standalone)
        current_tth_count = len(tth_pairs)*2 + len(tth_standalone)
        use_mwf = should_use_mwf(current_mwf_count, current_tth_count, 
                                 total_course_count, allow_unbalanced)
        
        # Decide pattern
        # For PathFit codes, assign strictly MW or TTH.
        # For all others, assign MWF or TTHS.
        if 'PathFit' in course.code or 'PATHFit' in course.code:
            pattern = 'MW' if use_mwf else 'TTH'
        else:
            pattern = 'MWF' if use_mwf else 'TTHS'
        
        is_mwf = pattern in ['MWF', 'MW']
        
        # Pick one available slot
        possible_slots = [ts for ts in time_slots if ts.is_available(is_mwf)]
        if not possible_slots:
            continue
        
        selected_slot = random.choice(possible_slots)
        selected_slot.mark_used(is_mwf)
        
        # Manage a simple "consecutive slot" rule for standalone courses
        day_key = 'mwf' if is_mwf else 'tth'
        consecutive_slot_count.setdefault(day_key, 0)
        
        # If we used 1 slot in a row, mark the next slot as used (simulate break)
        if consecutive_slot_count[day_key] == 1:
            idx = time_slots.index(selected_slot)
            if idx + 1 < len(time_slots):
                time_slots[idx + 1].mark_used(is_mwf)
            consecutive_slot_count[day_key] = 0
        else:
            consecutive_slot_count[day_key] += 1
        
        # Assign final scheduling info
        course.time = f"{selected_slot.start_time_str}-{selected_slot.end_time_str}"
        course.days = pattern
        course.room = select_room(course, available_rooms)
        course.time_obj = selected_slot.start_time
        course.end_time_obj = selected_slot.end_time
        
        if pattern in ['MWF', 'MW']:
            mwf_standalone.append(course)
        else:
            tth_standalone.append(course)
    
    # Combine MWF pairs + standalones
    mwf_courses = []
    for lec, lab in mwf_pairs:
        mwf_courses.extend([lec, lab])
    mwf_courses.extend(mwf_standalone)
    mwf_courses.sort(key=lambda x: x.time_obj)
    
    # Combine TTH pairs + standalones
    tth_courses = []
    for lec, lab in tth_pairs:
        tth_courses.extend([lec, lab])
    tth_courses.extend(tth_standalone)
    tth_courses.sort(key=lambda x: x.time_obj)
    
    # Final combined list
    return mwf_courses + tth_courses


############################################
#               MAIN APP UI                #
############################################

class SectionCreatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Subject Offering")
        self.root.geometry("460x800")
        self.root.configure(bg="#f0f0f0")
        self.df = None
        
        # Define programs and their available years
        self.program_years = {
            "BSIT(WebTech)": ["First", "Second", "Third"],
            "BSIT(Netsec)": ["First", "Second", "Third"],
            "BSIT(ERP)": ["First", "Second", "Third"],
            "BSCS": ["First", "Second", "Third"],
            "BSDA": ["First", "Second", "Third"],
            "BMMA": ["First", "Second", "Third"]
        }
        
        # Dict to store all section count widgets
        self.section_counts = {}
        self.trimester_var = tk.StringVar(value="First")

        self.setup_ui()

    def setup_ui(self):
        # Main container
        main_container = ttk.Frame(self.root, padding="20")
        main_container.grid(row=0, column=0, sticky="nsew")
        
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Style configuration
        style = ttk.Style()
        style.configure("Title.TLabel", font=("Helvetica", 16, "bold"))
        style.configure("Header.TLabel", font=("Helvetica", 12, "bold"))
        style.configure("Custom.TLabelframe", padding=15)
        style.configure("Action.TButton", padding=10)
        style.configure("Subheader.TLabel", font=("Helvetica", 10, "bold"))
        
        # Title
        title_frame = ttk.Frame(main_container)
        title_frame.grid(row=0, column=0, pady=(0, 20), sticky="ew")
        ttk.Label(title_frame, text="Subject Course Offering", style="Title.TLabel").pack()
        
        # File import frame
        import_frame = ttk.LabelFrame(main_container, text="Data Import", style="Custom.TLabelframe")
        import_frame.grid(row=1, column=0, pady=(0, 20), sticky="ew")
        
        ttk.Button(import_frame, text="Import CSV File", command=self.import_csv, style="Action.TButton").grid(row=0, column=0, padx=10)
        
        self.file_label = ttk.Label(import_frame, text="No file selected", font=("Helvetica", 10, "italic"))
        self.file_label.grid(row=0, column=1, padx=10)
        
        # Configuration frame
        config_frame = ttk.LabelFrame(main_container, text="Section Configuration", style="Custom.TLabelframe")
        config_frame.grid(row=2, column=0, pady=(0, 20), sticky="ew")
        
        # Trimester selection
        trim_frame = ttk.Frame(config_frame)
        trim_frame.grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="ew")
        
        ttk.Label(trim_frame, text="Trimester:", style="Header.TLabel").grid(row=0, column=0, padx=(0, 10))
        
        trimester_combo = ttk.Combobox(
            trim_frame, 
            textvariable=self.trimester_var,
            values=["First", "Second", "Third"],
            width=15,
            state="readonly"
        )
        trimester_combo.grid(row=0, column=1)
        
        # Program sections configuration
        sections_frame = ttk.Frame(config_frame)
        sections_frame.grid(row=1, column=0, sticky="ew", columnspan=2)
        
        ttk.Label(sections_frame, text="Number of Sections per Program", style="Header.TLabel").grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        left_frame = ttk.Frame(sections_frame)
        right_frame = ttk.Frame(sections_frame)
        left_frame.grid(row=1, column=0, padx=20, sticky="n")
        right_frame.grid(row=1, column=1, padx=20, sticky="n")
        
        left_programs = ["BSIT(WebTech)", "BSIT(Netsec)", "BSIT(ERP)"]
        self.create_program_spinboxes(left_programs, left_frame)
        
        right_programs = ["BSCS", "BSDA", "BMMA"]
        self.create_program_spinboxes(right_programs, right_frame)
        
        # Generate button
        generate_frame = ttk.Frame(main_container)
        generate_frame.grid(row=3, column=0, pady=20)
        
        ttk.Button(generate_frame, text="Generate and Export Sections",
                   command=self.generate_sections, style="Action.TButton", width=30).pack()

    def create_program_spinboxes(self, programs: List[str], parent_frame: ttk.Frame):
        current_row = 0
        for program in programs:
            ttk.Label(parent_frame, text=program, style="Subheader.TLabel").grid(
                row=current_row, column=0, sticky="w"
            )
            current_row += 1

            self.section_counts[program] = {}
            years = self.program_years[program]
            for year in years:
                current_row = self.add_year_spinbox(parent_frame, program, year, current_row)
            
            current_row += 1

    def add_year_spinbox(self, parent_frame: ttk.Frame, program: str, year: str, row_index: int) -> int:
        container = ttk.Frame(parent_frame)
        container.grid(row=row_index, column=0, pady=2, padx=20, sticky="w")
        
        ttk.Label(container, text=f"{year}:", width=8, anchor="e").grid(row=0, column=0, padx=(0, 10))
        spinbox = ttk.Spinbox(container, from_=1, to=10, width=5, justify="center")
        spinbox.grid(row=0, column=1)
        spinbox.set(3)
        
        self.section_counts[program][year] = spinbox
        return row_index + 1

    def import_csv(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if file_path:
            self.df = pd.read_csv(file_path)
            file_name = file_path.split("/")[-1]
            self.file_label.config(text=f"Selected file: {file_name}")
            messagebox.showinfo("Success", "CSV file imported successfully!")

    def create_section_courses(self, program: str, year: str, trimester: str) -> List[Course]:
        """
        Only pull courses from "BSIT" if:
          - program is one of [BSIT(WebTech), BSIT(Netsec), BSIT(ERP)]
          - AND we meet any of these (year, trimester) combos:
              (Third, Third), (First, Second), (First, First), (First, Third)
        Otherwise, pull from the specialization name in 'program'.
        """
        if self.df is None:
            return []
        
        # Conditions to fetch from "BSIT":
        allowed_combinations = {
            ("Third", "Third"),  # 3rd year, 3rd trimester
            ("First", "Second"), # 1st year, 2nd trimester
            ("First", "First"),  # 1st year, 1st trimester
            ("First", "Third")   # 1st year, 3rd trimester
        }
        
        if (
            program in ["BSIT(WebTech)", "BSIT(Netsec)", "BSIT(ERP)"] and
            (year, trimester) in allowed_combinations
        ):
            program_key = "BSIT"
        else:
            program_key = program
        
        filtered_df = self.df[
            (self.df['Program'] == program_key) &
            (self.df['Year_Level'] == year) &
            (self.df['Trimester'] == trimester)
        ]
        
        courses = []
        for _, row in filtered_df.iterrows():
            course = Course(
                row['Course_Code'],
                row['Description'],
                row['Units'],
                row['Time'],
                row['Days'],
                row['Room']
            )
            courses.append(course)
        
        return balance_and_shuffle_courses(courses, year, trimester)

    def generate_sections(self):
        if self.df is None:
            messagebox.showerror("Error", "Please import CSV file first")
            return

        wb = openpyxl.Workbook()
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

        year_mapping = {
            "First": "1",
            "Second": "2",
            "Third": "3"
        }

        for group_name in ['A', 'B']:
            ws = wb.create_sheet(f"Group {group_name}")
            row = 1

            for program, year_dict in self.section_counts.items():
                for year, spinbox in year_dict.items():
                    num_sections = int(spinbox.get())
                    if num_sections > 0:
                        for section_num in range(num_sections):
                            section_letter = chr(65 + section_num)
                            section_name = f"{year_mapping.get(year, '?')}{section_letter}"

                            ws.cell(row=row, column=1, 
                                    value=f"{program}, {year} Year, {self.trimester_var.get()} "
                                          f"Trimester Section {section_name}")
                            row += 1

                            headers = ["COURSE CODE", "DESCRIPTION", "UNITS", "TIME", "DAYS", "ROOM"]
                            for col_idx, header in enumerate(headers, 1):
                                ws.cell(row=row, column=col_idx, value=header)
                            row += 1

                            courses = self.create_section_courses(program, year, self.trimester_var.get())

                            mwf_courses = [c for c in courses if c.days in ['MWF', 'MW']]
                            tth_courses = [c for c in courses if c.days in ['TTH', 'TTHS']]

                            for c in mwf_courses:
                                ws.cell(row=row, column=1, value=c.code)
                                ws.cell(row=row, column=2, value=c.description)
                                ws.cell(row=row, column=3, value=c.units)
                                ws.cell(row=row, column=4, value=c.time)
                                ws.cell(row=row, column=5, value=c.days)
                                ws.cell(row=row, column=6, value=c.room)
                                row += 1

                            if mwf_courses and tth_courses:
                                ws.cell(row=row, column=1, value="----- End of MWF Schedule -----")
                                row += 1

                            for c in tth_courses:
                                ws.cell(row=row, column=1, value=c.code)
                                ws.cell(row=row, column=2, value=c.description)
                                ws.cell(row=row, column=3, value=c.units)
                                ws.cell(row=row, column=4, value=c.time)
                                ws.cell(row=row, column=5, value=c.days)
                                ws.cell(row=row, column=6, value=c.room)
                                row += 1

                            row += 1

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Success", "Sections generated and exported successfully!")


def main():
    root = tk.Tk()
    app = SectionCreatorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
